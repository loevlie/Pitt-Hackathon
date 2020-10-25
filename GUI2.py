from tkinter import *
from tkhtmlview import HTMLLabel
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from sklearn import preprocessing
import pickle
from openpyxl import Workbook, load_workbook
import datetime

tr=pd.read_csv('./healthcare/train_data.csv')
di=pd.read_csv('./healthcare/train_data_dictionary.csv')
df_train = pd.read_csv('./healthcare/train_data.csv',index_col='case_id')# ,index_col='case_id'
df_train=df_train.dropna()

#wb = Workbook()
#ws = wb.active
#ws.append(['time']+list(df_train.columns))
#wb.save("C:/Users/yashg/OneDrive/Desktop/CMU/Fall 2020/Misc/healthcare/entries.xlsx")

cats=['Hospital_type_code','Hospital_region_code','Department','Ward_Type','Ward_Facility_Code','Type of Admission','Severity of Illness','Age']
y=df_train['Stay']
d={}
for x in di['Column'][:-1]:
    ll=tr[x].unique()
    if len(ll)<50:
        d.update({x:ll})  
    else:
        d.update({x:'-'})

s=[]
filename='C:/Users/yashg/OneDrive/Desktop/CMU/Fall 2020/Misc/healthcare/rf.sav'
loaded_model = pickle.load(open(filename, 'rb'))

def press():
    df={dk[ix]:s[ix] for ix in range(len(s))}
    data=pd.DataFrame(df,index=[0])
    data=data.drop(['case_id','patientid','Admission_Deposit'],axis=1)
    for x in cats:
        le=preprocessing.LabelEncoder()
        le.fit(df_train[x])
        data[x]=le.transform(data[x])
    le2=preprocessing.LabelEncoder()
    le2.fit(y)
    y_pred=loaded_model.predict(data)
    ypred=le2.inverse_transform(y_pred)
    print(f'The patient should stay for {ypred[0]}')
    label=Label(root,text=f'The patient should ideally stay for \n{ypred[0]} days',borderwidth=2, relief="groove")
    label.grid(row=10,column=4)
    ct=datetime.datetime.now()
#    wb = load_workbook("C:/Users/yashg/OneDrive/Desktop/CMU/Fall 2020/Misc/healthcare/entries.xlsx")
#    ws = wb.active
#    ws.append([str(ct)]+s)
#    wb.save("C:/Users/yashg/OneDrive/Desktop/CMU/Fall 2020/Misc/healthcare/entries.xlsx")
    
def reset():
    global s
    s=[]

def create_dropdown(name,lis,root,mainframe):
    tkvar=StringVar(root)
#    l=sorted(lis)
    popupmenu=OptionMenu(mainframe,tkvar,*lis)
    tkvar.set('')
    tkvar.trace_add('write', lambda *args: s.append(tkvar.get()))
    w=Label(mainframe,text=name)
    return popupmenu,w,tkvar

#dk=sorted(d.keys())
dk=list(d.keys())
root = Tk()
root.title("Stay duration estimator")
mainframe = Frame(root)
mainframe.grid(column=0,row=0, sticky=(N,W,E,S) )
mainframe.columnconfigure(0, weight = 1)
mainframe.rowconfigure(0, weight = 1)

pp=[]
tkk=[]

for i,xx in enumerate(dk):
    w,x,tkv=create_dropdown(xx,d[xx],root,mainframe)
    tkk.append(tkv)
    pp.append([x,w])
count=0
for j,ls in enumerate(pp):
    if count<16:
        count=count+1
        x,w=ls
        x.grid(row=count,column=1)
        count=count+1
        w.grid(row=count,column=1)
    else:
        count=count+1
        x,w=ls
        x.grid(row=count-16,column=2)
        count=count+1
        w.grid(row=count-16,column=2)

B = Button(root, text ="Predict",command=press,font=("Times New Roman", 20)).grid(row=count+1,column=0)

#Rules='This serves as a modular way to predict how long a patient should stay \nin a hospital based on\ncertain prefixed parameters'
#w = Label(root, text=Rules,font=("Times New Roman", 10)).grid(row = count+2, column = 0)

root.mainloop()
